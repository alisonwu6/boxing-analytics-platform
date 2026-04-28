"""Excel export: sync report, kinematics, and fused signal."""

from typing import Optional

import openpyxl
import numpy as np

from ..core.imu import IMUData
from ..core.video import VideoKinematics
from ..core.sync import SyncResult, FusedDataset


def export_excel(kin: VideoKinematics,
                 imu_r: Optional[IMUData],
                 imu_l: Optional[IMUData],
                 sync: SyncResult,
                 fused: FusedDataset,
                 out_path: str):

    wb = openpyxl.Workbook()

    # Sheet 1 — Sync summary
    ws1 = wb.active
    ws1.title = "Sync Summary"
    _header_row(ws1, ["Parameter", "Value"])
    ws1.append(["offset_R (s)", round(float(sync.offset_R), 4)])
    ws1.append(["offset_L (s)", round(float(sync.offset_L), 4)])
    ws1.append(["inter_sensor (s)", round(float(sync.inter_sensor), 4)])
    ws1.append(["cc_offset_R (s)", round(float(sync.cc_offset_R), 4)])
    ws1.append(["cc_offset_L (s)", round(float(sync.cc_offset_L), 4)])
    ws1.append([])
    ws1.append(["Jump Events"])
    _header_row(ws1, ["#", "source", "time_s", "dur_s", "height_cm", "conf"])
    for i, j in enumerate(sync.video_jumps + sync.imu_r_jumps + sync.imu_l_jumps):
        ws1.append([i + 1, j.source, round(j.time_s, 4),
                    round(j.dur_s, 4), round(j.height_cm, 1), round(j.conf, 3)])
    _col_width(ws1)

    # Sheet 2 — Video kinematics
    ws2 = wb.create_sheet("Video Kinematics")
    hip = kin.lm_hip_y if kin.lm_hip_y is not None else np.full(len(kin.timestamps), np.nan)
    _header_row(ws2, ["time_s", "centroid_y", "acc_y", "jump_signal", "hip_y"])
    for i in range(len(kin.timestamps)):
        ws2.append([
            round(float(kin.timestamps[i]), 4),
            round(float(kin.centroid_y[i]), 5),
            round(float(kin.acc_y[i]), 5),
            round(float(kin.jump_signal[i]), 5),
            round(float(hip[i]), 5) if not np.isnan(hip[i]) else None,
        ])
    _col_width(ws2)

    # Sheet 3 — IMU-R
    if imu_r:
        ws3 = wb.create_sheet("IMU-R")
        _header_row(ws3, ["time_s", "acc_x", "acc_y", "acc_z", "magnitude"])
        for i in range(len(imu_r.timestamps)):
            ws3.append([
                round(float(imu_r.timestamps[i]), 4),
                round(float(imu_r.acc_x[i]), 5),
                round(float(imu_r.acc_y[i]), 5),
                round(float(imu_r.acc_z[i]), 5),
                round(float(imu_r.magnitude[i]), 5),
            ])
        _col_width(ws3)

    # Sheet 4 — IMU-L
    if imu_l:
        ws4 = wb.create_sheet("IMU-L")
        _header_row(ws4, ["time_s", "acc_x", "acc_y", "acc_z", "magnitude"])
        for i in range(len(imu_l.timestamps)):
            ws4.append([
                round(float(imu_l.timestamps[i]), 4),
                round(float(imu_l.acc_x[i]), 5),
                round(float(imu_l.acc_y[i]), 5),
                round(float(imu_l.acc_z[i]), 5),
                round(float(imu_l.magnitude[i]), 5),
            ])
        _col_width(ws4)

    # Sheet 5 — Fused 200 Hz
    ws5 = wb.create_sheet("Fused 200Hz")
    _header_row(ws5, ["time_s", "acc_y_R", "acc_y_L", "acc_y_vid", "source_R", "source_L"])
    for i in range(len(fused.times)):
        ws5.append([
            round(float(fused.times[i]), 4),
            round(float(fused.acc_y_R[i]), 5),
            round(float(fused.acc_y_L[i]), 5),
            round(float(fused.acc_y_vid[i]), 5),
            int(fused.source_R[i]),
            int(fused.source_L[i]),
        ])
    _col_width(ws5)

    wb.save(out_path)
    print(f"[ExcelExport] Saved → {out_path}")


def export_to_excel(video_kinematics, imu_R_synced, imu_L_synced,
                    fused_200hz, output_path: str):
    """DataFrame-based export (wrist pipeline)."""
    import math
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Wrist Kinematics"

    def _write_df(ws, df):
        if df is None or df.empty:
            return
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append([_safe(v) for v in row])

    def _safe(v):
        if v is None:
            return None
        try:
            if math.isnan(v):
                return None
        except TypeError:
            pass
        if isinstance(v, bool):
            return int(v)
        return v

    _write_df(ws1, video_kinematics)

    for title, df in [("IMU-R Synced", imu_R_synced),
                      ("IMU-L Synced", imu_L_synced),
                      ("Fused 200Hz",  fused_200hz)]:
        ws = wb.create_sheet(title)
        _write_df(ws, df)

    wb.save(output_path)
    print(f"[ExcelExport] Saved → {output_path}")


def _header_row(ws, headers: list):
    from openpyxl.styles import Font
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)


def _col_width(ws, width: int = 16):
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = width
