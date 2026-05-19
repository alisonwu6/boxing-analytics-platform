from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_excel(out_stem, fps, n_frames,
                 L_sp, R_sp, L_dsw, R_dsw, L_dsw_dot, R_dsw_dot, L_ea, R_ea,
                 punches, rejected_peaks):
    xl_path = str(Path(out_stem).with_suffix(".xlsx"))
    wb = openpyxl.Workbook()

    hdr_fill    = PatternFill("solid", fgColor="1F3864")
    hdr_font    = Font(bold=True, color="FFFFFF", size=10)
    hdr_align   = Alignment(horizontal="center", vertical="center")
    thin        = Side(border_style="thin", color="AAAAAA")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = hdr_align; c.border = cell_border
        ws.row_dimensions[1].height = 18

    def auto_width(ws, padding=4):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + padding

    def _v(x):
        return None if (isinstance(x, float) and np.isnan(x)) else round(float(x), 2)

    # sheet 1: per-frame signals
    ws1 = wb.active; ws1.title = "Signals"
    hdrs = ["frame", "time_s",
            "L_speed_px_s", "R_speed_px_s",
            "L_dsw_px", "R_dsw_px",
            "L_dsw_dot", "R_dsw_dot",
            "L_ea_deg", "R_ea_deg"]
    style_header(ws1, hdrs)

    detected_frames = set()
    for p in punches:
        detected_frames.update(range(p["start_fi"], p["end_fi"] + 1))
    rejected_frames = {r["peak_fi"] for r in rejected_peaks}
    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")

    for fi in range(n_frames):
        row = [fi, round(fi / fps, 4),
               _v(L_sp[fi]), _v(R_sp[fi]),
               _v(L_dsw[fi]), _v(R_dsw[fi]),
               _v(L_dsw_dot[fi]), _v(R_dsw_dot[fi]),
               _v(L_ea[fi]), _v(R_ea[fi])]
        ws1.append(row)
        fill = green_fill if fi in detected_frames else \
               yellow_fill if fi in rejected_frames else None
        if fill:
            for col in range(1, len(hdrs) + 1):
                ws1.cell(fi + 2, col).fill = fill
    auto_width(ws1)
    ws1.freeze_panes = "A2"

    # sheet 2: confirmed punches (jab or uppercut)
    ptype = punches[0].get("type", "jab") if punches else "jab"
    sheet_name = "Detected_Uppercuts" if ptype == "uppercut" else "Detected_Jabs"
    ws2 = wb.create_sheet(sheet_name)
    if ptype == "uppercut":
        hdrs2 = ["hand", "punch#",
                 "start_fi", "start_t_s",
                 "load_fi",  "load_t_s",
                 "peak_fi",  "peak_t_s",
                 "end_fi",   "end_t_s",
                 "load_min_wa", "peak_wa",
                 "peak_ea_deg", "peak_vy_px_s",
                 "peak_speed_px_s", "dur_s"]
        style_header(ws2, hdrs2)
        for p in punches:
            ws2.append([
                p["hand"], p["punch"],
                p["start_fi"], round(p["start_fi"] / fps, 3),
                p["load_fi"],  round(p["load_fi"]  / fps, 3),
                p["peak_fi"],  round(p["peak_fi"]  / fps, 3),
                p["end_fi"],   round(p["end_fi"]   / fps, 3),
                p["load_min_wa"], p["peak_wa"],
                p["peak_ea"], p["peak_vy"],
                p["peak_speed"], p["dur_s"],
            ])
    else:
        hdrs2 = ["hand", "punch#",
                 "start_fi", "start_t_s",
                 "peak_fi",  "peak_t_s",
                 "end_fi",   "end_t_s",
                 "ea_min_deg", "ea_peak_deg", "ea_max_drive_deg",
                 "dsw_start_px", "dsw_peak_px", "dsw_gain_px",
                 "peak_speed_px_s", "dur_s"]
        style_header(ws2, hdrs2)
        for p in punches:
            ws2.append([
                p["hand"], p["punch"],
                p["start_fi"], round(p["start_fi"] / fps, 3),
                p["peak_fi"],  round(p["peak_fi"]  / fps, 3),
                p["end_fi"],   round(p["end_fi"]   / fps, 3),
                p["ea_min"], p["ea_peak"], p.get("ea_max_drive", "-"),
                p["dsw_start"], p["dsw_peak"], p["dsw_gain"],
                p["peak_speed"], p["dur_s"],
            ])
    auto_width(ws2)

    # sheet 3: rejected peaks
    ws3 = wb.create_sheet("Rejected_Peaks")
    hdrs3 = ["hand", "peak_fi", "peak_t_s", "reason",
             "speed_px_s", "dsw_or_vy", "ea_or_wa"]
    style_header(ws3, hdrs3)
    for r in rejected_peaks:
        ws3.append([r["hand"], r["peak_fi"], r["peak_t"],
                    r["reason"], r["sp"],
                    r.get("dsw", r.get("vy", "-")),
                    r.get("ea",  r.get("wa",  "-"))])
    auto_width(ws3)

    wb.save(xl_path)
    print(f"Saved Excel: {xl_path}")
    return xl_path
